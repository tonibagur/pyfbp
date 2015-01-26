import traceback

class FlowElement(object):
    local_debug=False
    def __init__(self,connector=None,method=None,connector_largs=[],connector_kwargs={},name='undefined'):
        self.name=name
        self.connector=connector
        self.connector_kwargs=connector_kwargs
        self.connector_largs=connector_largs
        self.method=method
        if type(self)!=NullProcessor:
            self.out_objects=[NullProcessor()]
            self.out_in_objects=[NullProcessor()]
            self.error_objects=[Printer('unhandled error in object:'+name)]
    def compute_args(self,elem={}):
        new_largs=[x for x in self.connector_largs]
        new_kwargs=self.connector_kwargs.copy()
        for i,e in enumerate(new_largs):
            if type(e)==Ev:
                new_largs[i]=e.evaluate(elem)
        for k,v in new_kwargs.iteritems():
            if type(v)==Ev:
                new_kwargs[k]=v.evaluate(elem)
        return new_largs,new_kwargs
    def out(self,*out_objects):
        self.out_objects=out_objects
        return self
    def add_out(self,*out_objects):
        self.out_objects+=out_objects
        return self
    def error(self,*error_objects):
        self.error_objects=error_objects
        return self
    def add_error(self,*error_objects):
        self.error_objects+=error_objects
        return self
    def out_in(self,*out_in_objects):
        self.out_in_objects=out_in_objects
        return self
    def add_out_in(self,*out_in_objects):
        self.out_in_objects+=out_in_objects
        return self

    def out_elem(self,e,debug=False):
        for o in self.out_objects:
            o.process(e,debug)
    def out_in_elem(self,e,debug=False):
        for o in self.out_in_objects:
            o.process(e,debug)
    def error_elem(self,e,debug=False):
        for o in self.error_objects:
            o.process(e,debug)

class NullProcessor(FlowElement):
    def process(self,elem={},debug=False):
        return

class ListProcessor(FlowElement):
    def process(self,elem={},debug=False):
        if debug or self.local_debug:
            print "Debug mode ListProcessor:",self.name
            import pdb
            pdb.set_trace()
        largs,kwargs=self.compute_args(elem)
        for x in getattr(self.connector,self.method)(*largs,**kwargs):
            self.out_elem(x,debug)
            self.out_in_elem({'in':elem,'out':x},debug)

class ElementProcessor(FlowElement):
    def process(self,elem={},debug=False):
        try:
            if debug or self.local_debug:
                print "Debug mode ElementProcessor:",self.name
                import pdb
                pdb.set_trace()
            largs,kwargs=self.compute_args(elem)
            out=getattr(self.connector,self.method)(*largs,**kwargs)
            self.out_elem(out,debug)
            self.out_in_elem({'in':elem,'out':out},debug)
        except:
            error=traceback.format_exc()
            elem['error']=error
            self.error_elem(elem,debug)



class ElementFilter(FlowElement):
    def __init__(self,filter=None,debug=False,name='undefined'):
        self.name=name
        self.filter=filter
        self.out_objects=[NullProcessor()]
        self.out_in_objects=[NullProcessor()]
        self.error_objects=[Printer('unhandled error in object:'+name)]
        self.out_rejected_objects=[NullProcessor()]

    def out_rejected(self,*out_rejected_objects):
        self.out_rejected_objects=out_rejected_objects
        return self

    def out_rejected_elem(self,e,debug=False):
        for o in self.out_rejected_objects:
            o.process(e,debug)

    def process(self,elem={},debug=False):
        if debug or self.local_debug:
            print "Debug mode ElementFilter:",self.name
            import pdb
            pdb.set_trace()
        if not self.filter or self.filter.evaluate(elem):
            self.out_elem(elem,debug)
            self.out_in_elem(elem,debug)
        else:
            self.out_rejected_elem(elem,debug)

class ListAcumulator(FlowElement):
    def __init__(self,name='undefined'):
        self.name=name
        self.list=[]
        self.out_objects=[NullProcessor()]
        self.out_in_objects=[NullProcessor()]
        self.error_objects=[Printer('unhandled error in object:'+name)]
    def process(self,elem={},debug=False):
        if debug or self.local_debug:
            print "Debug mode ListAcumulator:",self.name
            import pdb
            pdb.set_trace()
        self.list.append(elem)
        self.out_elem(elem,debug)
        self.out_in_elem(elem,debug)


    def count(self):
        return len(self.list)

    def __getitem__(self,i):
        return self.list[i]

class BreakPoint(FlowElement):
    def __init__(self,**debug_vars):
        self.list=[]
        self.out_objects=[NullProcessor()]
        self.error_objects=[Printer('unhandled error in object:'+name)]
        self.out_in_objects=[NullProcessor()]
        if 'name' in debug_vars:
            self.name=debug_vars['name']
            del debug_vars['name']
        else:
            self.name='undefined'
        self.debug_vars=debug_vars
    def process(self,elem={},debug=False):
        #TODO, Veure si ara es necessari tenint en compte que debug_vars pot ser util
        print "Breakpoint:",self.name
        import pdb
        pdb.set_trace()
        self.out_elem(elem)
        self.out_in_elem(elem)


class Printer(FlowElement):
    def __init__(self,prefix=''):
        self.prefix=prefix
        self.out_objects=[NullProcessor()]
        self.out_in_objects=[NullProcessor()]
        self.error_objects=[NullProcessor()]
    def process(self,elem,debug=False):
        if debug or self.local_debug:
            print "Debug mode Printer:",self.prefix
            import pdb
            pdb.set_trace()
        print self.prefix,elem
        self.out_elem(elem,debug)
        self.out_in_elem(elem,debug)

class ElementTransformer(FlowElement):
    def __init__(self,name='undefined'):
        self.out_objects=[NullProcessor()]
        self.out_in_objects=[NullProcessor()]
        self.error_objects=[Printer('unhandled error in object:'+name)]
        self.name=name
    def process(self,elem,debug=False):
        try:
            if debug or self.local_debug:
                print "Debug mode Transformer",self.name
                import pdb
                pdb.set_trace()
            e2=self.transform(elem)
            self.out_elem(e2,debug)
            self.out_in_elem({'in':elem,'out':e2},debug)
        except:
            error=traceback.format_exc()
            elem['error']=error
            self.error_elem(elem)
    def transform(self,elem):
        return elem

class Ev(object):
    def __init__(self,expression):
        self.expression=expression

    def evaluate(self,elem):
        return eval(self.expression)

class ExcelReader(object):
    def __init__(self,file_name='excel.xlsx',sheet_name='Sheet1',read_header=True):
        self.sheet_name=sheet_name
        self.file_name=file_name
        self.read_header=read_header
        self.cols=None
        from openpyxl import load_workbook
        self.wb = load_workbook(filename = self.file_name, use_iterators = True)
        self.ws = self.wb.get_sheet_by_name(self.sheet_name) # ws is now an IterableWorksheet
    
    def read_rows(self):
        for row in self.ws.iter_rows(): # it brings a new method: iter_rows()
            if not self.cols:
                if self.read_header:
                    self.cols=[cell.value for cell in row]
                else:
                    self.cols=['col'+str(i) for i in range(len(row))]
                    yield self.build_row(row)
            else:
                yield self.build_row(row)
    def build_row(self,row):
        return dict([(self.cols[i],row[i].value) for i,c in enumerate(row)])

class ExcelOutput(FlowElement):
    def __init__(self,file_name='excel.xlsx',columns=None,name='undefined',write_header=True):
        self.file_name=file_name
        self.columns=columns
        self.name=name
        self.out_objects=[NullProcessor()]
        self.out_in_objects=[NullProcessor()]
        self.error_objects=[Printer('unhandled error in object:'+name)]
        from openpyxl import Workbook
        self.wb = Workbook(write_only=True)
        self.ws = self.wb.create_sheet()
        self.write_header=write_header
        self.header_written=False


    def process(self,elem,debug=False):
        try:
            if debug or self.local_debug:
                print "Debug mode ExcelOutput",self.name
                import pdb
                pdb.set_trace()
            if not self.columns:
                self.columns=elem.keys()
            if self.write_header and not self.header_written:
                self.header_written=True
                self.ws.append([c for c in self.columns])
            self.ws.append([elem[x] for x in self.columns ])


        except:
            error=traceback.format_exc()
            elem['error']=error
            self.error_elem(elem)

    def save(self):
        self.wb.save(self.file_name)


class CSVOutput(FlowElement):
    def __init__(self,file_name='output.csv',columns=None,name='undefined',write_header=True,separator=";"):
        self.file_name=file_name
        self.columns=columns
        self.name=name
        self.out_objects=[NullProcessor()]
        self.out_in_objects=[NullProcessor()]
        self.error_objects=[Printer('unhandled error in object:'+name)]
        self.file = open(self.file_name,'w')
        self.write_header=write_header
        self.header_written=False
        self.separator=separator


    def process(self,elem,debug=False):
        try:
            if debug or self.local_debug:
                print "Debug mode CSVOutput",self.name
                import pdb
                pdb.set_trace()
            if not self.columns:
                self.columns=elem.keys()
            if self.write_header and not self.header_written:
                self.header_written=True
                self.file.write(self.separator.join([c for c in self.columns])+'\n')
            self.file.write(self.separator.join([str(elem[x]) for x in self.columns ])+'\n')
        except:
            error=traceback.format_exc()
            elem['error']=error
            self.error_elem(elem)

    def save(self):
        self.file.close()

def test():
    return True

if __name__=='__main__':
    '''class Test(object):
        def get_list(self):
            return [{'val':i, 'val2':i+1} for i in range(100)]
    csv=CSVOutput(columns=['val','val2'])
    ListProcessor(Test(),'get_list').out(
        Printer(),
        csv,
    ).process(debug=False)'''
    '''xls=ExcelOutput(columns=['val','val2'])
    ListProcessor(Test(),'get_list').out(
        Printer(),
        xls,
    ).process()
    xls.save()'''
    '''xlsr=ExcelReader(file_name='excel.xlsx',sheet_name='Sheet1',read_header=True)
    ListProcessor(xlsr,'read_rows').out(
        Printer()
    ).process()'''

    ef=ElementFilter(Ev('test()')).out(Printer()).process({})




